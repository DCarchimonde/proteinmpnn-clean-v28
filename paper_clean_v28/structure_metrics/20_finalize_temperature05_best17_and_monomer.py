#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Create the final complex-T0.5 plus monomer all-metrics workbook."""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional

import numpy as np
import pandas as pd


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent.parent
DEFAULT_RUN_DIR = ROOT / "paper_clean_v28_outputs" / "temperature_0.5_best17"
DEFAULT_MODEL_SUMMARY = ROOT / "paper_clean_v28_outputs" / "monomer_clean" / "summary.json"
EXPECTED_COMPLEX_ROWS = 17
EXPECTED_MONOMER_ROWS = 151


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge complex best17 and monomer metrics into one workbook."
    )
    parser.add_argument("--run_dir", default=str(DEFAULT_RUN_DIR))
    parser.add_argument("--model_summary", default=str(DEFAULT_MODEL_SUMMARY))
    return parser.parse_args()


def finite(value: object) -> bool:
    try:
        return math.isfinite(float(value))
    except Exception:
        return False


def numeric(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame:
        return pd.Series(dtype=float)
    return pd.to_numeric(frame[column], errors="coerce")


def count_status(frame: pd.DataFrame, column: str, expected: str = "ok") -> int:
    if column not in frame:
        return 0
    return int(frame[column].astype(str).str.lower().eq(expected.lower()).sum())


def median_or_nan(frame: pd.DataFrame, column: str) -> float:
    return float(numeric(frame, column).median())


def mean_or_nan(frame: pd.DataFrame, column: str) -> float:
    return float(numeric(frame, column).mean())


def add_metric(
    rows: List[Dict[str, object]],
    *,
    scope: str,
    category: str,
    metric: str,
    value: object,
    unit: str = "",
    n_available: object = "",
    status: str = "computed",
    definition: str = "",
) -> None:
    rows.append(
        {
            "scope": scope,
            "category": category,
            "metric": metric,
            "value": value,
            "unit": unit,
            "n_available": n_available,
            "status": status,
            "definition": definition,
        }
    )


def build_complex_summary(frame: pd.DataFrame) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    n = len(frame)
    global_rmsd = numeric(frame, "global_complex_ca_rmsd")
    cyclic_rmsd = numeric(
        frame,
        "cyclic_peptide_ca_rmsd_after_global_complex_alignment_best_forward_cyclic_shift",
    )
    joint = (
        global_rmsd.lt(3) & cyclic_rmsd.lt(3)
    )
    add_metric(
        rows,
        scope="Complex_T0.5_best17",
        category="Panel",
        metric="selected_structures",
        value=n,
        unit="structures",
        n_available=n,
        definition="One RMSD-best structure per target at generation temperature 0.5.",
    )
    add_metric(
        rows,
        scope="Complex_T0.5_best17",
        category="Sequence",
        metric="median_natural_aa_recovery",
        value=median_or_nan(frame, "natural_aa_recovery"),
        unit="fraction",
        n_available=int(numeric(frame, "natural_aa_recovery").notna().sum()),
    )
    for metric, series, definition in [
        (
            "median_global_complex_ca_rmsd",
            global_rmsd,
            "Whole-complex CA RMSD after one global PyMOL alignment.",
        ),
        (
            "median_cyclic_peptide_ca_rmsd_best_forward_shift",
            cyclic_rmsd,
            "Complete final-chain peptide CA RMSD in the unchanged global-alignment frame; minimum over forward cyclic shifts.",
        ),
    ]:
        add_metric(
            rows,
            scope="Complex_T0.5_best17",
            category="RMSD",
            metric=metric,
            value=float(series.median()),
            unit="angstrom",
            n_available=int(series.notna().sum()),
            definition=definition,
        )
    for metric, count in [
        ("global_complex_ca_rmsd_lt_3", int(global_rmsd.lt(3).sum())),
        ("cyclic_peptide_ca_rmsd_lt_3", int(cyclic_rmsd.lt(3).sum())),
        ("joint_global_and_cyclic_ca_rmsd_lt_3", int(joint.sum())),
    ]:
        add_metric(
            rows,
            scope="Complex_T0.5_best17",
            category="RMSD",
            metric=metric,
            value=count,
            unit=f"of {n}",
            n_available=n,
        )

    confidence_columns = [
        ("median_highfold_ca_bfactor_plddt", "highfold_ca_bfactor_mean", "score"),
        ("median_highfold_comment_plddt", "highfold_plddt_comment", "score"),
        ("median_highfold_comment_ptm", "highfold_ptm_comment", "score"),
        ("median_peptide_chain_plddt", "peptide_chain_plddt", "score"),
        ("median_peptide_chain_iptm", "peptide_chain_iptm", "score"),
        ("median_peptide_receptor_iptm", "peptide_receptor_iptm_mean", "score"),
        (
            "median_peptide_receptor_inter_pae",
            "peptide_receptor_inter_pae_mean",
            "angstrom",
        ),
    ]
    for metric, column, unit in confidence_columns:
        values = numeric(frame, column)
        add_metric(
            rows,
            scope="Complex_T0.5_best17",
            category="Confidence",
            metric=metric,
            value=float(values.median()),
            unit=unit,
            n_available=int(values.notna().sum()),
            status="computed_available_cases",
        )

    structure_columns = [
        (
            "median_legacy_peptide_ca_rmsd_after_receptor_fit",
            "legacy_peptide_ca_rmsd_after_receptor_fit",
        ),
        (
            "median_legacy_peptide_backbone_rmsd_after_receptor_fit",
            "legacy_peptide_backbone_rmsd_after_receptor_fit",
        ),
        (
            "median_legacy_peptide_ca_rmsd_self_superposed",
            "legacy_peptide_ca_rmsd_self_superposed",
        ),
        (
            "median_methyl_ca_rmsd_after_peptide_self_fit",
            "methyl_ca_rmsd_after_peptide_self_fit",
        ),
        (
            "median_nonmethyl_ca_rmsd_after_peptide_self_fit",
            "nonmethyl_ca_rmsd_after_peptide_self_fit",
        ),
        (
            "median_methyl_backbone_rmsd_after_peptide_self_fit",
            "methyl_backbone_rmsd_after_peptide_self_fit",
        ),
        (
            "median_nonmethyl_backbone_rmsd_after_peptide_self_fit",
            "nonmethyl_backbone_rmsd_after_peptide_self_fit",
        ),
    ]
    for metric, column in structure_columns:
        values = numeric(frame, column)
        add_metric(
            rows,
            scope="Complex_T0.5_best17",
            category="Position_RMSD",
            metric=metric,
            value=float(values.median()),
            unit="angstrom",
            n_available=int(values.notna().sum()),
        )

    permeability = numeric(frame, "permeability_pred")
    add_metric(
        rows,
        scope="Complex_T0.5_best17",
        category="Permeability",
        metric="median_permeability_pred",
        value=float(permeability.median()),
        n_available=int(permeability.notna().sum()),
    )

    energy_columns = [
        (
            "median_complex_score_per_residue",
            "rosetta_complex_score_per_residue",
            "REU/residue",
        ),
        (
            "median_receptor_subset_score_per_residue",
            "rosetta_receptor_subset_score_per_residue",
            "REU/residue",
        ),
        (
            "median_peptide_subset_score_per_residue",
            "rosetta_peptide_subset_score_per_residue",
            "REU/residue",
        ),
        (
            "median_cross_interface_energy_fixed_pose",
            "rosetta_cross_interface_energy_fixed_pose",
            "REU",
        ),
        (
            "median_cross_interface_energy_per_peptide_residue",
            "rosetta_cross_interface_energy_per_peptide_residue",
            "REU/residue",
        ),
    ]
    for metric, column, unit in energy_columns:
        values = numeric(frame, column)
        add_metric(
            rows,
            scope="Complex_T0.5_best17",
            category="PyRosetta",
            metric=metric,
            value=float(values.median()),
            unit=unit,
            n_available=int(values.notna().sum()),
            definition="Fixed-pose naturalized ref2015 score; not kcal/mol.",
        )
    interface = numeric(frame, "rosetta_cross_interface_energy_fixed_pose")
    add_metric(
        rows,
        scope="Complex_T0.5_best17",
        category="PyRosetta",
        metric="cross_interface_energy_lt_0",
        value=int(interface.lt(0).sum()),
        unit=f"of {n}",
        n_available=int(interface.notna().sum()),
    )

    unavailable = [
        (
            "Diversity",
            "within_target_tm_diversity_1_minus_tm",
            "One structure per target leaves zero within-target pairs.",
        ),
        (
            "Energy_delta",
            "energy_success_vs_native",
            "Native peptide reference energy was not computed by the fixed-pose workflow.",
        ),
        (
            "Energy_delta",
            "energy_stability_vs_native",
            "Native complex reference energy was not computed by the fixed-pose workflow.",
        ),
        (
            "Binding_site",
            "binding_site_recovery_ratio",
            "No validated binding-site definition/cutoff was established.",
        ),
    ]
    for category, metric, reason in unavailable:
        add_metric(
            rows,
            scope="Complex_T0.5_best17",
            category=category,
            metric=metric,
            value=np.nan,
            status="not_estimable",
            definition=reason,
        )
    return rows


def build_monomer_summary(frame: pd.DataFrame) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    n = len(frame)
    add_metric(
        rows,
        scope="Monomer_151",
        category="Panel",
        metric="samples",
        value=n,
        unit="samples",
        n_available=n,
    )
    add_metric(
        rows,
        scope="Monomer_151",
        category="Panel",
        metric="explicit_methyl_structure_pairs_available",
        value=count_status(frame, "explicit_methyl_structure_status", "ok"),
        unit=f"of {n}",
        n_available=n,
    )
    recovery = numeric(frame, "natural_aa_recovery_fixed_order")
    add_metric(
        rows,
        scope="Monomer_151",
        category="Sequence",
        metric="mean_natural_aa_recovery_fixed_order",
        value=float(recovery.mean()),
        unit="fraction",
        n_available=int(recovery.notna().sum()),
    )

    structure_columns = [
        (
            "median_naturalized_ca_rmsd_best_forward_shift",
            "naturalized_ca_rmsd_best_forward_cyclic_shift",
            "angstrom",
        ),
        (
            "median_naturalized_backbone_rmsd_after_ca_fit",
            "naturalized_backbone_rmsd_after_ca_fit_best_forward_cyclic_shift",
            "angstrom",
        ),
        (
            "median_e2e_methyl_ca_rmsd",
            "naturalized_e2e_methyl_ca_rmsd_after_best_cyclic_ca_fit",
            "angstrom",
        ),
        (
            "median_e2e_nonmethyl_ca_rmsd",
            "naturalized_e2e_nonmethyl_ca_rmsd_after_best_cyclic_ca_fit",
            "angstrom",
        ),
        (
            "median_explicit_methyl_pair_ca_rmsd",
            "explicit_methyl_ca_rmsd_best_forward_cyclic_shift",
            "angstrom",
        ),
        (
            "median_tm_score_best_forward_shift",
            "naturalized_tm_score_symmetric_best_forward_cyclic_shift",
            "score",
        ),
        (
            "median_diversity_1_minus_tm",
            "naturalized_diversity_1_minus_tm_best_forward_cyclic_shift",
            "score",
        ),
    ]
    for metric, column, unit in structure_columns:
        values = numeric(frame, column)
        add_metric(
            rows,
            scope="Monomer_151",
            category="Structure",
            metric=metric,
            value=float(values.median()),
            unit=unit,
            n_available=int(values.notna().sum()),
            definition=(
                "HighFold reference-sequence versus e2e-design prediction; "
                "not experimental native validation."
            ),
        )
    rmsd = numeric(frame, "naturalized_ca_rmsd_best_forward_cyclic_shift")
    for threshold in (2, 3, 5):
        add_metric(
            rows,
            scope="Monomer_151",
            category="Structure",
            metric=f"naturalized_ca_rmsd_lt_{threshold}",
            value=int(rmsd.lt(threshold).sum()),
            unit=f"of {n}",
            n_available=int(rmsd.notna().sum()),
        )

    confidence_columns = [
        (
            "median_reference_ca_bfactor_plddt",
            "reference_naturalized_ca_bfactor_mean",
        ),
        (
            "median_e2e_ca_bfactor_plddt",
            "e2e_naturalized_ca_bfactor_mean",
        ),
        ("median_reference_comment_ptm", "reference_naturalized_comment_ptm"),
        ("median_e2e_comment_ptm", "e2e_naturalized_comment_ptm"),
    ]
    for metric, column in confidence_columns:
        values = numeric(frame, column)
        add_metric(
            rows,
            scope="Monomer_151",
            category="Confidence",
            metric=metric,
            value=float(values.median()),
            unit="score",
            n_available=int(values.notna().sum()),
            status="computed_available_cases",
        )

    for role in ("reference", "e2e"):
        values = numeric(frame, f"{role}_permeability_pred")
        add_metric(
            rows,
            scope="Monomer_151",
            category="Permeability",
            metric=f"median_{role}_permeability_pred",
            value=float(values.median()),
            n_available=int(values.notna().sum()),
            status=(
                "computed"
                if int(values.notna().sum()) == n
                else "partial_or_not_found"
            ),
        )
    permeability_delta = numeric(frame, "permeability_delta_e2e_minus_reference")
    permeability_higher = numeric(frame, "e2e_permeability_gt_reference")
    add_metric(
        rows,
        scope="Monomer_151",
        category="Permeability",
        metric="median_permeability_delta_e2e_minus_reference",
        value=float(permeability_delta.median()),
        n_available=int(permeability_delta.notna().sum()),
    )
    add_metric(
        rows,
        scope="Monomer_151",
        category="Permeability",
        metric="e2e_permeability_gt_reference",
        value=int(permeability_higher.sum()),
        unit=f"of {int(permeability_higher.notna().sum())}",
        n_available=int(permeability_higher.notna().sum()),
    )

    energy_columns = [
        (
            "median_reference_rosetta_score_per_residue",
            "reference_rosetta_score_per_residue",
        ),
        (
            "median_e2e_rosetta_score_per_residue",
            "e2e_rosetta_score_per_residue",
        ),
        (
            "median_rosetta_score_per_residue_delta_e2e_minus_reference",
            "rosetta_score_per_residue_delta_e2e_minus_reference",
        ),
    ]
    for metric, column in energy_columns:
        values = numeric(frame, column)
        add_metric(
            rows,
            scope="Monomer_151",
            category="PyRosetta",
            metric=metric,
            value=float(values.median()),
            unit="REU/residue",
            n_available=int(values.notna().sum()),
            definition="Fixed-pose naturalized ref2015 score; not kcal/mol.",
        )
    lower = numeric(
        frame, "e2e_lower_rosetta_score_per_residue_than_reference"
    )
    add_metric(
        rows,
        scope="Monomer_151",
        category="PyRosetta",
        metric="e2e_lower_rosetta_score_per_residue_than_reference",
        value=int(lower.sum()),
        unit=f"of {int(lower.notna().sum())}",
        n_available=int(lower.notna().sum()),
        definition="Descriptive fixed-pose comparison, not experimental stability.",
    )

    unavailable = [
        (
            "Interface",
            "cross_interface_energy",
            "A single-chain monomer has no receptor-peptide interface.",
        ),
        (
            "Confidence",
            "iptm",
            "ipTM is an inter-chain confidence metric and is not applicable.",
        ),
        (
            "Confidence",
            "inter_pae",
            "Inter-chain PAE is not applicable to a single-chain monomer.",
        ),
        (
            "Binding_pose",
            "receptor_fit_binding_pose_rmsd",
            "A monomer has no receptor frame.",
        ),
        (
            "Binding_site",
            "binding_site_recovery_ratio",
            "A monomer has no receptor-defined binding site in this workflow.",
        ),
        (
            "RMSD",
            "all_atom_rmsd",
            "Reference and e2e sequences differ; no validated side-chain atom mapping.",
        ),
    ]
    for category, metric, reason in unavailable:
        add_metric(
            rows,
            scope="Monomer_151",
            category=category,
            metric=metric,
            value=np.nan,
            status="not_applicable",
            definition=reason,
        )
    return rows


def flatten_model_summary(path: Path) -> pd.DataFrame:
    payload = json.loads(path.read_text(encoding="utf-8"))
    strict = next(
        (
            item
            for item in payload
            if item.get("input_mode") == "strict_naturalized_input"
        ),
        None,
    )
    if strict is None:
        raise ValueError("strict_naturalized_input row missing from monomer summary")
    rows: List[Dict[str, object]] = []
    for metric in [
        "n_selected_positions",
        "n_methyl_positive",
        "n_methyl_negative",
        "true_methyl_rate",
        "base_recovery",
        "known_sequence_auc",
        "end_to_end_auc",
    ]:
        rows.append(
            {
                "input_mode": "strict_naturalized_input",
                "task": "overall",
                "metric": metric,
                "value": strict.get(metric),
            }
        )
    for source_key, task in [
        ("known_sequence_best_by_f1", "known_sequence_methylation"),
        ("end_to_end_best_by_f1", "end_to_end_methylation"),
    ]:
        metrics = strict.get(source_key, {})
        for metric in [
            "threshold",
            "accuracy",
            "precision",
            "recall",
            "f1",
            "false_positive_rate",
            "pred_methyl_rate",
            "tp",
            "tn",
            "fp",
            "fn",
        ]:
            rows.append(
                {
                    "input_mode": "strict_naturalized_input",
                    "task": task,
                    "metric": metric,
                    "value": metrics.get(metric),
                }
            )
    return pd.DataFrame(rows)


def add_model_summary_metrics(
    summary_rows: List[Dict[str, object]], model_metrics: pd.DataFrame
) -> None:
    for _, row in model_metrics.iterrows():
        add_metric(
            summary_rows,
            scope="Monomer_model",
            category=str(row["task"]),
            metric=str(row["metric"]),
            value=row["value"],
            status="computed_clean_evaluation",
            definition="Clean evaluation, strict_naturalized_input.",
        )


def build_quality_checks(
    complex_frame: pd.DataFrame,
    monomer_frame: pd.DataFrame,
) -> pd.DataFrame:
    checks: List[Dict[str, object]] = []

    def add(
        scope: str,
        check: str,
        expected: object,
        observed: object,
        passed: bool,
        required: bool = True,
        note: str = "",
    ):
        checks.append(
            {
                "scope": scope,
                "check": check,
                "expected": expected,
                "observed": observed,
                "required": int(required),
                "status": "PASS" if passed else ("WARN" if not required else "FAIL"),
                "note": note,
            }
        )

    add(
        "Complex",
        "rows",
        EXPECTED_COMPLEX_ROWS,
        len(complex_frame),
        len(complex_frame) == EXPECTED_COMPLEX_ROWS,
    )
    add(
        "Complex",
        "unique_targets",
        EXPECTED_COMPLEX_ROWS,
        complex_frame.get("target_name", pd.Series(dtype=str)).nunique(),
        complex_frame.get("target_name", pd.Series(dtype=str)).nunique()
        == EXPECTED_COMPLEX_ROWS,
    )
    temperatures = numeric(complex_frame, "temperature")
    add(
        "Complex",
        "temperature_0.5_rows",
        EXPECTED_COMPLEX_ROWS,
        int(temperatures.eq(0.5).sum()),
        int(temperatures.eq(0.5).sum()) == EXPECTED_COMPLEX_ROWS,
    )
    for check, column in [
        ("global_rmsd_available", "global_complex_ca_rmsd"),
        (
            "cyclic_peptide_rmsd_available",
            "cyclic_peptide_ca_rmsd_after_global_complex_alignment_best_forward_cyclic_shift",
        ),
        ("permeability_available", "permeability_pred"),
        ("complex_energy_available", "rosetta_complex_score_per_residue"),
        (
            "cross_interface_energy_available",
            "rosetta_cross_interface_energy_fixed_pose",
        ),
    ]:
        observed = int(numeric(complex_frame, column).notna().sum())
        add(
            "Complex",
            check,
            EXPECTED_COMPLEX_ROWS,
            observed,
            observed == EXPECTED_COMPLEX_ROWS,
        )

    add(
        "Monomer",
        "rows",
        EXPECTED_MONOMER_ROWS,
        len(monomer_frame),
        len(monomer_frame) == EXPECTED_MONOMER_ROWS,
    )
    add(
        "Monomer",
        "unique_samples",
        EXPECTED_MONOMER_ROWS,
        monomer_frame.get("sample_name", pd.Series(dtype=str)).nunique(),
        monomer_frame.get("sample_name", pd.Series(dtype=str)).nunique()
        == EXPECTED_MONOMER_ROWS,
    )
    for check, column, expected_status in [
        ("naturalized_structure_pairs", "naturalized_structure_status", "ok"),
        ("tm_pairs", "naturalized_tm_status", "ok"),
        ("paired_pyrosetta_energy", "paired_energy_status", "ok"),
    ]:
        observed = count_status(monomer_frame, column, expected_status)
        add(
            "Monomer",
            check,
            EXPECTED_MONOMER_ROWS,
            observed,
            observed == EXPECTED_MONOMER_ROWS,
        )
    for check, column in [
        ("reference_naturalized_pdbs", "variant_2_present"),
        ("e2e_naturalized_pdbs", "variant_4_present"),
    ]:
        observed = int(numeric(monomer_frame, column).sum())
        add(
            "Monomer",
            check,
            EXPECTED_MONOMER_ROWS,
            observed,
            observed == EXPECTED_MONOMER_ROWS,
        )
    for check, column in [
        ("reference_permeability_available", "reference_permeability_pred"),
        ("e2e_permeability_available", "e2e_permeability_pred"),
    ]:
        observed = int(numeric(monomer_frame, column).notna().sum())
        add(
            "Monomer",
            check,
            EXPECTED_MONOMER_ROWS,
            observed,
            observed == EXPECTED_MONOMER_ROWS,
            required=False,
            note="Auto-discovered from local monomer permeability CSVs; missing values remain NA.",
        )
    return pd.DataFrame(checks)


def definitions_table() -> pd.DataFrame:
    rows = [
        {
            "object": "Complex",
            "metric_family": "Primary RMSD",
            "definition": "One global complex alignment, then complete final-chain peptide CA RMSD under every forward cyclic shift; no peptide-only second fit.",
            "applicability": "17 temperature-0.5 RMSD-best complex structures",
        },
        {
            "object": "Complex",
            "metric_family": "Cross-interface energy",
            "definition": "E_complex - E_receptor_subset - E_peptide_subset in the intact fixed pose.",
            "applicability": "Complex only; REU, not binding free energy",
        },
        {
            "object": "Monomer",
            "metric_family": "Primary RMSD",
            "definition": "CA self-superposition of e2e design to reference-sequence prediction; minimum over all forward cyclic shifts.",
            "applicability": "151 naturalized reference/e2e pairs",
        },
        {
            "object": "Monomer",
            "metric_family": "TM/diversity",
            "definition": "Best-forward-shift symmetric TM-score; diversity = 1 - TM.",
            "applicability": "151 naturalized reference/e2e pairs; descriptive for short peptides",
        },
        {
            "object": "Monomer",
            "metric_family": "Explicit methyl sensitivity",
            "definition": "Same CA/backbone comparison for PDB variants 1 and 3 when both exist.",
            "applicability": "110 expected available pairs; missing variants remain NA",
        },
        {
            "object": "Both",
            "metric_family": "PyRosetta",
            "definition": "Fixed-pose naturalized ref2015 score with rama_prepro, omega, p_aa_pp disabled.",
            "applicability": "REU; no relaxation and no explicit N-methyl chemistry",
        },
        {
            "object": "Monomer",
            "metric_family": "Not applicable",
            "definition": "Cross-interface energy, ipTM, inter-PAE, receptor-fit pose RMSD, and BSR require multiple chains/a receptor.",
            "applicability": "Recorded as NA, never zero",
        },
    ]
    return pd.DataFrame(rows)


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            maximum = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells[: min(len(column_cells), 250)]
            )
            sheet.column_dimensions[letter].width = min(max(maximum + 2, 10), 45)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
    workbook.save(path)


def main() -> int:
    args = parse_args()
    run_dir = Path(args.run_dir).resolve()
    monomer_dir = run_dir / "monomer"
    model_summary_path = Path(args.model_summary).resolve()

    complex_path = run_dir / "temperature05_best17_all_metrics.csv"
    monomer_structure_path = monomer_dir / "monomer_structure_metrics_by_sample.csv"
    monomer_energy_path = monomer_dir / "monomer_pyrosetta_energy_by_sample.csv"
    for path in [
        complex_path,
        monomer_structure_path,
        monomer_energy_path,
        model_summary_path,
    ]:
        if not path.is_file():
            raise FileNotFoundError(path)

    complex_frame = pd.read_csv(complex_path)
    monomer_structure = pd.read_csv(monomer_structure_path)
    monomer_energy = pd.read_csv(monomer_energy_path)
    monomer_frame = monomer_structure.merge(
        monomer_energy,
        how="left",
        on="sample_name",
        validate="one_to_one",
        suffixes=("", "_energy"),
    )
    model_metrics = flatten_model_summary(model_summary_path)

    complex_frame.insert(0, "object_scope", "complex_temperature_0.5_best17")
    monomer_frame.insert(0, "object_scope", "monomer_reference_vs_e2e")
    monomer_frame["cross_interface_energy_status"] = (
        "not_applicable_single_chain_monomer"
    )
    monomer_frame["iptm_status"] = "not_applicable_single_chain_monomer"
    monomer_frame["inter_pae_status"] = "not_applicable_single_chain_monomer"
    monomer_frame["receptor_fit_binding_pose_rmsd_status"] = (
        "not_applicable_no_receptor"
    )
    monomer_frame["binding_site_recovery_status"] = (
        "not_applicable_no_receptor_defined_binding_site"
    )

    summary_rows = build_complex_summary(complex_frame)
    summary_rows.extend(build_monomer_summary(monomer_frame))
    add_model_summary_metrics(summary_rows, model_metrics)
    metric_summary = pd.DataFrame(summary_rows)
    quality = build_quality_checks(complex_frame, monomer_frame)
    definitions = definitions_table()

    required_failures = quality[
        quality["required"].eq(1) & quality["status"].eq("FAIL")
    ]
    warnings = quality[quality["status"].eq("WARN")]
    quality_pass = len(required_failures) == 0

    workbook_path = run_dir / "temperature05_best17_and_monomer_all_metrics.xlsx"
    summary_path = run_dir / "temperature05_best17_and_monomer_metric_summary.csv"
    monomer_path = run_dir / "monomer_all_metrics.csv"
    quality_path = run_dir / "temperature05_best17_and_monomer_quality_gate.csv"
    report_path = run_dir / "temperature05_best17_and_monomer_report.txt"

    metric_summary.to_csv(summary_path, index=False, encoding="utf-8-sig")
    monomer_frame.to_csv(monomer_path, index=False, encoding="utf-8-sig")
    quality.to_csv(quality_path, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        metric_summary.to_excel(writer, sheet_name="Metric_summary", index=False)
        complex_frame.to_excel(writer, sheet_name="Complex_best17", index=False)
        monomer_frame.to_excel(writer, sheet_name="Monomer_151", index=False)
        model_metrics.to_excel(writer, sheet_name="Monomer_model", index=False)
        quality.to_excel(writer, sheet_name="Quality_gate", index=False)
        definitions.to_excel(writer, sheet_name="Definitions", index=False)
    style_workbook(workbook_path)

    lines = [
        "===== COMPLEX T0.5 BEST17 + MONOMER 151 FINAL REPORT =====",
        f"Complex rows: {len(complex_frame)}/{EXPECTED_COMPLEX_ROWS}",
        f"Complex unique targets: {complex_frame['target_name'].nunique()}",
        f"Monomer rows: {len(monomer_frame)}/{EXPECTED_MONOMER_ROWS}",
        f"Monomer unique samples: {monomer_frame['sample_name'].nunique()}",
        f"Required quality checks failed: {len(required_failures)}",
        f"Non-blocking warnings: {len(warnings)}",
        "",
        "Workbook sheets:",
        "- Metric_summary",
        "- Complex_best17",
        "- Monomer_151",
        "- Monomer_model",
        "- Quality_gate",
        "- Definitions",
        "",
        "Scope notes:",
        "- Complex rows are restricted to the exact 17 RMSD-best structures at temperature 0.5.",
        "- Monomer primary structure rows use all 151 complete naturalized reference/e2e pairs.",
        "- Explicit-methyl monomer structure metrics are a coverage-limited sensitivity subset.",
        "- Inapplicable or scientifically non-estimable fields remain NA, never zero.",
        "",
        f"QUALITY GATE: {'PASS' if quality_pass else 'FAIL'}",
        f"PROBLEMS: {len(required_failures)}",
        f"WARNINGS: {len(warnings)}",
    ]
    if len(required_failures):
        lines.extend(
            ["", "Required failures:"]
            + [
                f"- {row['scope']} / {row['check']}: "
                f"expected={row['expected']}, observed={row['observed']}"
                for _, row in required_failures.iterrows()
            ]
        )
    if len(warnings):
        lines.extend(
            ["", "Warnings:"]
            + [
                f"- {row['scope']} / {row['check']}: "
                f"expected={row['expected']}, observed={row['observed']}"
                for _, row in warnings.iterrows()
            ]
        )
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    (run_dir / "temperature05_best17_and_monomer_manifest.json").write_text(
        json.dumps(
            {
                "quality_gate": "PASS" if quality_pass else "FAIL",
                "complex_rows": len(complex_frame),
                "monomer_rows": len(monomer_frame),
                "required_failures": int(len(required_failures)),
                "warnings": int(len(warnings)),
                "workbook": str(workbook_path),
                "metric_summary_csv": str(summary_path),
                "monomer_all_metrics_csv": str(monomer_path),
                "quality_gate_csv": str(quality_path),
                "report": str(report_path),
            },
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )

    print("===== FINAL COMPLEX + MONOMER WORKBOOK COMPLETE =====")
    print(f"complex rows: {len(complex_frame)}")
    print(f"monomer rows: {len(monomer_frame)}")
    print(f"quality gate: {'PASS' if quality_pass else 'FAIL'}")
    print(f"workbook: {workbook_path}")
    print(f"report: {report_path}")
    return 0 if quality_pass else 1


if __name__ == "__main__":
    sys.exit(main())
