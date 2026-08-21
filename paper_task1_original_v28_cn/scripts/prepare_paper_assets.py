#!/usr/bin/env python3
"""Prepare traceable Task-1 paper tables, supplementary files, and figures.

This script does not rerun the neural model or any structure predictor.  It only
reads already-audited outputs and converts them into publication assets.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import shutil
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PAPER = ROOT / "paper_task1_original_v28_cn"
DATA = PAPER / "data"
FIG = PAPER / "figures"
SUPP = PAPER / "supplementary"
TASK1 = ROOT / "work" / "task1"
V9 = ROOT / "work" / "v9" / "cyclic_stability_v9_1700"
OUT = V9 / "paper_clean_v28_outputs"


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return int(value)
    return value


def sheet_rows(path: Path, sheet: str, header_marker: str | None = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    if header_marker is None:
        start = 0
    else:
        start = next(i for i, row in enumerate(rows) if row and row[0] == header_marker)
    header = [str(v) if v is not None else "" for v in rows[start]]
    width = len(header)
    body = []
    for row in rows[start + 1 :]:
        values = [clean_cell(v) for v in row[:width]]
        if not any(v != "" for v in values):
            continue
        body.append(values)
    return header, body


def write_csv(path: Path, header, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerows(rows)


def copy_file(src: Path, dst: Path):
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)


def copy_or_reuse(src: Path, dst: Path):
    """Copy a local attachment when present, otherwise retain the bundled copy."""
    if src.exists():
        copy_file(src, dst)
    elif not dst.exists():
        raise FileNotFoundError(f"Missing both source attachment and bundled copy: {dst}")


def prefer_source(src: Path, bundled: Path) -> Path:
    """Use the archived upstream source when present, otherwise the bundled copy."""
    if src.exists():
        return src
    if bundled.exists():
        return bundled
    raise FileNotFoundError(f"Missing both upstream source and bundled copy: {src} / {bundled}")


def latex_escape(text: str) -> str:
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    return "".join(replacements.get(ch, ch) for ch in str(text))


def fmt_float(value, digits=3):
    if value in (None, ""):
        return "--"
    return f"{float(value):.{digits}f}"


def prepare_workbook_exports():
    main_book = prefer_source(
        TASK1 / "Task1_六复合物_T0.5_论文主表与图表数据.xlsx",
        SUPP / "Task1_six_complexes_paper_tables.xlsx",
    )
    audit_book = prefer_source(
        TASK1 / "Task1_六复合物_T0.5_逐序列完整审计.xlsx",
        SUPP / "Task1_six_complexes_476_sequence_audit.xlsx",
    )

    main_specs = {
        "Pub_Table1": "target_name",
        "Pub_Table2": "target_name",
        "Pub_Table3": "target_name",
        "Table4_Model": "dataset",
        "Table5_Availability": "metric",
        "Source_Index": "category",
    }
    for sheet, marker in main_specs.items():
        header, rows = sheet_rows(main_book, sheet, marker)
        write_csv(DATA / f"{sheet}.csv", header, rows)

    audit_specs = [
        "Candidates_476",
        "Pass_LT3_16",
        "Pass_LT5_101",
        "RMSD_Summary",
        "Position_Frequency",
        "RMSD_Sensitivity",
        "Model_ByResidue",
        "Threshold_Curves",
        "QC",
        "Sources",
    ]
    for sheet in audit_specs:
        header, rows = sheet_rows(audit_book, sheet)
        write_csv(DATA / f"{sheet}.csv", header, rows)

    main_dst = SUPP / "Task1_six_complexes_paper_tables.xlsx"
    audit_dst = SUPP / "Task1_six_complexes_476_sequence_audit.xlsx"
    if main_book.resolve() != main_dst.resolve():
        copy_file(main_book, main_dst)
    if audit_book.resolve() != audit_dst.resolve():
        copy_file(audit_book, audit_dst)


def prepare_normalized_sequence_metric_tables():
    """Split presentation-oriented workbook exports into standard one-table CSVs."""
    threshold = pd.read_csv(DATA / "Threshold_Curves.csv", skiprows=2)
    assert len(threshold) == 20
    assert set(threshold["task"]) == {"known_natural_sequence", "end_to_end"}
    threshold.to_csv(
        DATA / "Monomer_Threshold_Curves_Corrected.csv",
        index=False,
        encoding="utf-8-sig",
    )

    main_metrics = pd.read_csv(DATA / "Table4_Model.csv", nrows=4)
    assert len(main_metrics) == 4
    assert set(main_metrics["positions"].astype(int)) == {75, 1505}
    main_metrics.to_csv(
        DATA / "Sequence_Main_Metrics_Corrected.csv",
        index=False,
        encoding="utf-8-sig",
    )

    historical = pd.read_csv(DATA / "Table4_Model.csv", skiprows=6)
    assert len(historical) == 2
    historical.to_csv(
        DATA / "Monomer_Historical_Labels_2.csv",
        index=False,
        encoding="utf-8-sig",
    )


def prepare_existing_figures_and_sources():
    for src in sorted((TASK1 / "figures").glob("Fig*.*")):
        copy_file(src, FIG / src.name)

    uploads = ROOT / "upload"
    copy_or_reuse(uploads / "7704196a-7fcb-4e7f-bc26-493b7a86e45b.png", SUPP / "historical_padding_e2e_table.png")
    copy_or_reuse(uploads / "427011f379f4d2421cc7e732cd0750e5.png", SUPP / "historical_padding_binary_table.png")
    copy_or_reuse(uploads / "3cf2ffed-e7b4-41f1-a498-ab73e9025f9f.png", SUPP / "requested_metric_outline.png")


def prepare_clean_existing_outputs():
    copies = {
        OUT / "generated_fasta_clean_auto_single" / "summary_by_temperature.csv": DATA / "all17_generation_by_temperature.csv",
        OUT / "complex_native_clean" / "summary.json": DATA / "all17_native_complex_summary.json",
        OUT / "complex_native_clean" / "threshold_metrics.csv": DATA / "all17_native_complex_threshold_metrics.csv",
        OUT / "structure_metrics" / "complex_rmsd_summary_by_temperature.csv": DATA / "best85_rmsd_by_temperature.csv",
        OUT / "structure_metrics" / "complex_methylation_site_rmsd_summary_by_temperature.csv": DATA / "best85_methyl_site_rmsd_by_temperature.csv",
        OUT / "structure_metrics" / "complex_structural_diversity_tm_pairwise.csv": DATA / "best85_diversity_pairwise_170.csv",
        OUT / "permeability" / "complex_permeability_best85.csv": DATA / "best85_permeability_85.csv",
        OUT / "structure_metrics" / "complex_pyrosetta_energy_naturalized_best85.csv": DATA / "best85_naturalized_fixed_pose_energy_85.csv",
        OUT / "structure_metrics" / "complex_rmsd_metrics.csv": DATA / "best85_structure_metrics_85.csv",
        OUT / "monomer_clean" / "position_predictions.csv": DATA / "Monomer_RawPredictions_4515.csv",
        V9 / "v9_inputs" / "test_serine_provenance_corrected.jsonl": DATA / "monomer_test_serine_provenance_corrected.jsonl",
    }
    for src, dst in copies.items():
        copy_or_reuse(src, dst)


def prepare_corrected_monomer_position_table():
    """Join frozen probabilities to the corrected test labels used in the paper.

    The upstream inference CSV intentionally preserves the historical 323-positive
    target tokens. The paper metrics use the Ser-provenance-corrected JSONL
    instead, so we publish a separate, unambiguous 1505-row main-analysis table.
    """
    raw_path = prefer_source(
        OUT / "monomer_clean" / "position_predictions.csv",
        DATA / "Monomer_RawPredictions_4515.csv",
    )
    raw = pd.read_csv(raw_path)
    strict = raw.loc[raw["input_mode"] == "strict_naturalized_input"].copy()

    corrected_sequences = {}
    corrected_jsonl = prefer_source(
        V9 / "v9_inputs" / "test_serine_provenance_corrected.jsonl",
        DATA / "monomer_test_serine_provenance_corrected.jsonl",
    )
    with corrected_jsonl.open(encoding="utf-8") as handle:
        for line in handle:
            record = json.loads(line)
            assert record["seq"] == record["seq_chain_A"]
            corrected_sequences[record["name"]] = record["seq"]

    assert len(corrected_sequences) == 151
    corrected_tokens = []
    for row in strict.itertuples(index=False):
        seq = corrected_sequences[row.sample_name]
        pos = int(row.position_in_model)
        assert 0 <= pos < len(seq)
        corrected_tokens.append(seq[pos])

    strict = strict.rename(
        columns={
            "target_token": "target_token_legacy",
            "target_token_index": "target_token_index_legacy",
            "is_methyl_true": "is_methyl_true_legacy",
        }
    )
    strict["target_token_corrected"] = corrected_tokens
    token_index_pairs = strict[["target_token_legacy", "target_token_index_legacy"]].drop_duplicates()
    assert token_index_pairs.groupby("target_token_legacy")["target_token_index_legacy"].nunique().max() == 1
    index_by_token = token_index_pairs.set_index("target_token_legacy")["target_token_index_legacy"].to_dict()
    index_by_token.update({token: idx for idx, token in enumerate("ACDEFGHIKLMNPQRSTVWY")})
    strict["target_token_index_corrected"] = [int(index_by_token[token]) for token in corrected_tokens]
    strict["is_methyl_true_corrected"] = [int(token.islower()) for token in corrected_tokens]
    strict["serine_provenance_label_changed"] = (
        (strict["is_methyl_true_legacy"] == 1)
        & (strict["is_methyl_true_corrected"] == 0)
        & (strict["true_base_token"] == "S")
    ).astype(int)
    strict["pred_methyl_known_strict_gt_0.6"] = (
        strict["prob_methyl_known_sequence"] > 0.6
    ).astype(int)
    strict["pred_methyl_end_to_end_strict_gt_0.6"] = (
        strict["prob_methyl_end_to_end"] > 0.6
    ).astype(int)
    strict["known_methyl_state_correct"] = (
        strict["pred_methyl_known_strict_gt_0.6"] == strict["is_methyl_true_corrected"]
    ).astype(int)
    strict["end_to_end_methyl_state_correct"] = (
        strict["pred_methyl_end_to_end_strict_gt_0.6"] == strict["is_methyl_true_corrected"]
    ).astype(int)
    strict["end_to_end_exact_methyl_residue_recovered"] = (
        (strict["is_methyl_true_corrected"] == 1)
        & (strict["base_correct"] == 1)
        & (strict["pred_methyl_end_to_end_strict_gt_0.6"] == 1)
    ).astype(int)
    strict["end_to_end_extended_token_correct"] = (
        (strict["base_correct"] == 1)
        & (strict["end_to_end_methyl_state_correct"] == 1)
    ).astype(int)
    strict.insert(
        2,
        "source_panel",
        np.where(strict["sample_name"].str.startswith("Me_"), "Rosetta-2023", "Rosetta-2025"),
    )

    assert len(strict) == 1505
    assert strict["sample_name"].nunique() == 151
    assert not strict.duplicated(["sample_name", "position_in_model"]).any()
    assert int(strict["is_methyl_true_legacy"].sum()) == 323
    assert int(strict["is_methyl_true_corrected"].sum()) == 261
    assert int(strict["serine_provenance_label_changed"].sum()) == 62
    changed = strict.loc[strict["serine_provenance_label_changed"] == 1]
    assert set(changed["target_token_legacy"]) == {"s"}
    assert set(changed["target_token_corrected"]) == {"S"}
    assert set(changed["target_token_index_legacy"]) == {34}
    assert set(changed["target_token_index_corrected"]) == {15}
    assert int(strict["base_correct"].sum()) == 242
    assert int(strict["pred_methyl_known_strict_gt_0.6"].sum()) == 189
    assert int(strict["pred_methyl_end_to_end_strict_gt_0.6"].sum()) == 125
    assert int(strict["end_to_end_exact_methyl_residue_recovered"].sum()) == 20
    assert int(strict["end_to_end_extended_token_correct"].sum()) == 224

    strict.to_csv(DATA / "Monomer_Corrected_1505.csv", index=False, encoding="utf-8-sig")


def prepare_historical_padding_audit():
    thresholds = [0.30, 0.40, 0.50, 0.60, 0.70, 0.80, 0.90, 0.95, 0.98, 0.99]
    pred_counts_b16 = [275, 268, 265, 256, 253, 240, 234, 213, 176, 58]
    exact_counts_b16 = [745, 743, 743, 742, 742, 739, 738, 735, 731, 712]
    rows = []
    for t, pred, correct in zip(thresholds, pred_counts_b16, exact_counts_b16):
        rows.append([t, 1947, 1505, 442, 323, pred, pred / 1947, correct, correct / 1947])
    write_csv(
        DATA / "historical_padding_batch16_reconstruction.csv",
        ["threshold", "tensor_positions", "real_positions", "padding_positions", "legacy_positives", "predicted_positive_n", "predicted_positive_rate", "legacy_exact_correct_n", "legacy_exact_accuracy"],
        rows,
    )

    confusion = [
        (0.30, 246, 843, 725, 77),
        (0.40, 178, 446, 1122, 145),
        (0.50, 128, 175, 1393, 195),
        (0.60, 90, 38, 1530, 233),
        (0.70, 77, 6, 1562, 246),
        (0.80, 58, 0, 1568, 265),
        (0.90, 4, 0, 1568, 319),
        (0.95, 0, 0, 1568, 323),
        (0.98, 0, 0, 1568, 323),
        (0.99, 0, 0, 1568, 323),
    ]
    out = []
    for t, tp, fp, tn, fn in confusion:
        n = tp + fp + tn + fn
        precision = tp / (tp + fp) if tp + fp else math.nan
        recall = tp / (tp + fn) if tp + fn else 0.0
        f1 = 2 * precision * recall / (precision + recall) if precision and precision + recall else 0.0
        out.append([t, n, 1505, 386, tp, fp, tn, fn, (tp + tn) / n, precision, recall, f1])
    write_csv(
        DATA / "historical_padding_batch8_reconstruction.csv",
        ["threshold", "tensor_positions", "real_positions", "padding_positions", "tp", "fp", "tn", "fn", "accuracy", "precision", "recall", "f1"],
        out,
    )


def prepare_compact_candidate_lists():
    df = pd.read_csv(DATA / "Pass_LT5_101.csv")
    cols = [
        "audit_row_id",
        "target_name",
        "design_seq",
        "methyl_positions_1based",
        "global_rmsd_A",
        "cyclic_rmsd_best_forward_A",
        "joint_rmsd_category",
        "pdb_file",
    ]
    df[cols].to_csv(DATA / "pass_lt5_compact_101.csv", index=False, encoding="utf-8-sig")
    df3 = df[df["joint_lt3"].astype(bool)]
    df3[cols].to_csv(DATA / "pass_lt3_compact_16.csv", index=False, encoding="utf-8-sig")

    rows_path = PAPER / "tables" / "pass_lt5_rows.tex"
    with rows_path.open("w", encoding="utf-8") as handle:
        for i, row in df.reset_index(drop=True).iterrows():
            category = r"$<3$" if bool(row["joint_lt3"]) else r"$3\text{--}<5$"
            handle.write(
                f"{i + 1} & {latex_escape(row['target_name'])} & "
                f"\\seq{{{latex_escape(row['design_seq'])}}} & "
                f"{latex_escape(row['methyl_positions_1based'])} & "
                f"{float(row['global_rmsd_A']):.3f} & "
                f"{float(row['cyclic_rmsd_best_forward_A']):.3f} & {category} \\\\\n"
            )


def prepare_best85_summary_and_figure():
    rmsd = pd.read_csv(prefer_source(
        OUT / "structure_metrics" / "complex_rmsd_metrics.csv",
        DATA / "best85_structure_metrics_85.csv",
    ))
    div = pd.read_csv(prefer_source(
        OUT / "structure_metrics" / "complex_structural_diversity_tm_pairwise.csv",
        DATA / "best85_diversity_pairwise_170.csv",
    ))
    perm = pd.read_csv(prefer_source(
        OUT / "permeability" / "complex_permeability_best85.csv",
        DATA / "best85_permeability_85.csv",
    ))
    energy = pd.read_csv(prefer_source(
        OUT / "structure_metrics" / "complex_pyrosetta_energy_naturalized_best85.csv",
        DATA / "best85_naturalized_fixed_pose_energy_85.csv",
    ))

    summary = [
        ["Peptide CA RMSD after receptor fit", "Å", len(rmsd), rmsd["peptide_ca_rmsd_after_receptor_fit"].mean(), rmsd["peptide_ca_rmsd_after_receptor_fit"].median(), "binding-pose metric; old best85 definition"],
        ["Peptide backbone RMSD after receptor fit", "Å", len(rmsd), rmsd["peptide_backbone_rmsd_after_receptor_fit"].mean(), rmsd["peptide_backbone_rmsd_after_receptor_fit"].median(), "binding-pose metric; old best85 definition"],
        ["Peptide CA RMSD after peptide self-fit", "Å", len(rmsd), rmsd["peptide_ca_rmsd_self_superposed"].mean(), rmsd["peptide_ca_rmsd_self_superposed"].median(), "internal shape only"],
        ["Peptide backbone RMSD after peptide self-fit", "Å", len(rmsd), rmsd["peptide_backbone_rmsd_self_superposed"].mean(), rmsd["peptide_backbone_rmsd_self_superposed"].median(), "internal shape only"],
        ["Symmetric TM-score", "unitless", len(div), div["tm_score_symmetric_mean"].mean(), div["tm_score_symmetric_mean"].median(), "170 within-target temperature pairs"],
        ["Diversity (1-TM)", "unitless", len(div), div["diversity_1_minus_tm"].mean(), div["diversity_1_minus_tm"].median(), "oracle best85; internal-shape diversity"],
        ["Permeability prediction", "model unit", len(perm), perm["permeability_pred"].mean(), perm["permeability_pred"].median(), "prediction without experimental labels"],
        ["Complex score per residue", "REU", len(energy), energy["rosetta_complex_score_per_residue"].mean(), energy["rosetta_complex_score_per_residue"].median(), "naturalized fixed pose"],
        ["Cross-interface score", "REU", len(energy), energy["rosetta_cross_interface_energy_fixed_pose"].mean(), energy["rosetta_cross_interface_energy_fixed_pose"].median(), "not binding free energy"],
    ]
    write_csv(DATA / "best85_support_summary.csv", ["metric", "unit", "n", "mean", "median", "interpretation"], summary)

    plt.style.use("seaborn-v0_8-whitegrid")
    fig, axes = plt.subplots(2, 2, figsize=(12.5, 9.0), constrained_layout=True)

    ax = axes[0, 0]
    ax.hist(div["diversity_1_minus_tm"].dropna(), bins=np.linspace(0, 1, 16), color="#0b84a5", edgecolor="white")
    ax.axvline(div["diversity_1_minus_tm"].mean(), color="#d95f02", linestyle="--", linewidth=2, label=f"Mean={div['diversity_1_minus_tm'].mean():.3f}")
    ax.set(xlabel="Diversity (1 - symmetric TM-score)", ylabel="Pair count", title="Structural diversity (170 pairs)")
    ax.legend(frameon=False)

    ax = axes[0, 1]
    temps = sorted(perm["temperature"].unique())
    vals = [perm.loc[perm["temperature"] == t, "permeability_log10"].dropna() for t in temps]
    ax.boxplot(vals, tick_labels=[f"{t:g}" for t in temps], patch_artist=True, boxprops={"facecolor": "#8dd3c7"}, medianprops={"color": "#d95f02", "linewidth": 2})
    ax.set(xlabel="Sampling temperature", ylabel="log10 permeability prediction", title="Predicted permeability (best85)")

    ax = axes[1, 0]
    vals = [energy.loc[energy["temperature"] == t, "rosetta_cross_interface_energy_fixed_pose"].dropna() for t in temps]
    ax.boxplot(vals, tick_labels=[f"{t:g}" for t in temps], patch_artist=True, boxprops={"facecolor": "#fdb462"}, medianprops={"color": "#1b9e77", "linewidth": 2}, showfliers=False)
    ax.axhline(0, color="#444444", linestyle="--", linewidth=1.5)
    neg = int((energy["rosetta_cross_interface_energy_fixed_pose"] < 0).sum())
    ax.set(xlabel="Sampling temperature", ylabel="Fixed-pose cross-interface score (REU)", title=f"Naturalized fixed-pose score ({neg}/85 < 0)")

    ax = axes[1, 1]
    labels = ["CA B-factor\nproxy", "COMMENT\npLDDT", "Peptide\npLDDT", "ipTM / PAE", "Permeability", "PyRosetta"]
    counts = [int(rmsd["highfold_ca_bfactor_available"].sum()), int(rmsd["highfold_plddt"].notna().sum()), int(rmsd["peptide_chain_plddt"].notna().sum()), int(rmsd["peptide_receptor_iptm_mean"].notna().sum()), len(perm), len(energy)]
    bars = ax.bar(labels, counts, color=["#80b1d3", "#bebada", "#bebada", "#bebada", "#8dd3c7", "#fdb462"])
    ax.set_ylim(0, 92)
    ax.set(ylabel="Available designs (of 85)", title="Metric completeness")
    ax.tick_params(axis="x", labelrotation=20)
    for bar, count in zip(bars, counts):
        ax.text(bar.get_x() + bar.get_width() / 2, count + 1.5, str(count), ha="center", va="bottom", fontsize=10)

    for label, ax in zip("ABCD", axes.flat):
        ax.text(-0.12, 1.08, label, transform=ax.transAxes, fontsize=16, fontweight="bold", va="top")

    fig.savefig(FIG / "Fig6_best85_exploratory_support.pdf", bbox_inches="tight")
    fig.savefig(FIG / "Fig6_best85_exploratory_support.png", dpi=600, bbox_inches="tight")
    fig.savefig(FIG / "Fig6_best85_exploratory_support.svg", bbox_inches="tight")
    plt.close(fig)


def prepare_manifest():
    checkpoint_source = prefer_source(ROOT / "frankenstein_v28.pt", V9 / "frankenstein_v28.pt")
    files = {
        "canonical_checkpoint": ("frankenstein_v28.pt", checkpoint_source),
        "serine_corrected_test_labels": (
            "paper_task1_original_v28_cn/data/monomer_test_serine_provenance_corrected.jsonl",
            DATA / "monomer_test_serine_provenance_corrected.jsonl",
        ),
        "task1_paper_workbook": (
            "paper_task1_original_v28_cn/supplementary/Task1_six_complexes_paper_tables.xlsx",
            SUPP / "Task1_six_complexes_paper_tables.xlsx",
        ),
        "task1_sequence_audit_workbook": (
            "paper_task1_original_v28_cn/supplementary/Task1_six_complexes_476_sequence_audit.xlsx",
            SUPP / "Task1_six_complexes_476_sequence_audit.xlsx",
        ),
        "monomer_corrected_position_table": (
            "paper_task1_original_v28_cn/data/Monomer_Corrected_1505.csv",
            DATA / "Monomer_Corrected_1505.csv",
        ),
        "monomer_raw_probability_table": (
            "paper_task1_original_v28_cn/data/Monomer_RawPredictions_4515.csv",
            DATA / "Monomer_RawPredictions_4515.csv",
        ),
        "monomer_pdb_provenance_audit": (
            "paper_task1_original_v28_cn/data/Monomer_PDB_Provenance_751.csv",
            DATA / "Monomer_PDB_Provenance_751.csv",
        ),
        "monomer_exact_overlap_audit": (
            "paper_task1_original_v28_cn/data/Monomer_Train_Test_Exact_Overlap_Audit.json",
            DATA / "Monomer_Train_Test_Exact_Overlap_Audit.json",
        ),
    }

    def artifact_entry(path: Path):
        return {
            "path": str(path.relative_to(ROOT)),
            "size_bytes": path.stat().st_size,
            "sha256": sha256(path),
        }

    bundled_artifacts = {
        str(path.relative_to(PAPER)): artifact_entry(path)
        for directory in (DATA, FIG, SUPP, PAPER / "fonts")
        for path in sorted(directory.glob("*"))
        if path.is_file()
    }
    manifest = {
        "paper_scope": "Task1 only: canonical original checkpoint, monomer evaluation, six non-3AV T=0.5 primary analysis, and explicitly exploratory all-17 best85 support",
        "repository": "DCarchimonde/proteinmpnn-clean-v28",
        "base_commit": "c9718e2a433a51f187606676dfedd11eeb6fc8a5",
        "legacy_source_repository": "DCarchimonde/ProteinMPNN",
        "legacy_source_commit": "28dff152d83623dfb322480413b7dc889f8537a4",
        "legacy_dataset_git_blobs": {
            "combined_751": "768d7863ac1ad58ab8935471d9bce3d8132d1631",
            "train_600": "720b7e65d4bf7c415ab6529e5561449332706809",
            "test_151": "bd18c23881cb51051691199de484a3ed33274861",
        },
        "monomer_dataset_audit": {
            "combined_751_sha256": "f35a91c7e1a0f4fbbda31dd1da6bdd7674810aad9b982e4e67c4a3738ae1ee89",
            "train_600_sha256": "a6064401bb3c8bda1522c41780e37fa744f19d8ab9cbf72b4cece05689f7b0f8",
            "test_151_sha256": "3a5d1fcfcb94a0a9850e7ab715e955effdea7ece524534513e2cb54514ece3f5",
            "raw_pdb_audit": "751/751 EXPDTA THEORETICAL MODEL and MODEL GENERATED BY ROSETTA",
            "rosetta_2023_panel": "406 Me_ records; Rosetta 2023.35+release.23439d3",
            "rosetta_2025_panel": "345 pdb_ records; Rosetta 2025.25+release.a0cefad01b",
            "baker_33_status": "not found; planning screenshot conflicts with frozen raw PDB provenance",
            "row_level_audit": "paper_task1_original_v28_cn/data/Monomer_PDB_Provenance_751.csv",
            "audit_script": "paper_task1_original_v28_cn/scripts/audit_monomer_pdb_provenance.py",
            "exact_overlap_report": "paper_task1_original_v28_cn/data/Monomer_Train_Test_Exact_Overlap_Audit.json",
            "homology_clustering_status": "not performed",
        },
        "files": {
            name: {"path": display_path, "size_bytes": path.stat().st_size, "sha256": sha256(path)}
            for name, (display_path, path) in files.items()
        },
        "bundled_artifacts": bundled_artifacts,
        "reproducibility": {
            "latex_build": "self-contained after clone; run build.ps1 or build.sh",
            "asset_refresh": "prepare_paper_assets.py prefers archived maintainer-workspace sources and otherwise reuses bundled paper copies",
            "legacy_pdb_reaudit": "requires a checkout of DCarchimonde/ProteinMPNN at the frozen legacy commit",
        },
        "metric_boundaries": {
            "monomer_structure_metrics": "not available; no result table was found",
            "baker_33_rosetta_400_plus_300": "historical outline conflicts with frozen data: all 751 audited monomer PDBs are Rosetta theoretical models (406 from Rosetta-2023 and 345 from Rosetta-2025)",
            "six_complex_primary_rmsd": "one global complex CA alignment followed by complete final-chain peptide CA RMSD under best forward cyclic shift; no peptide-only refit",
            "best85_selection": "oracle best-of-N by native sequence recovery; exploratory and selection-conditioned",
        },
    }
    (PAPER / "SOURCE_MANIFEST.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main():
    for directory in (DATA, FIG, SUPP, PAPER / "tables"):
        directory.mkdir(parents=True, exist_ok=True)
    prepare_workbook_exports()
    prepare_normalized_sequence_metric_tables()
    prepare_existing_figures_and_sources()
    prepare_clean_existing_outputs()
    prepare_corrected_monomer_position_table()
    prepare_historical_padding_audit()
    prepare_compact_candidate_lists()
    prepare_best85_summary_and_figure()
    prepare_manifest()
    print("Prepared paper assets in", PAPER)


if __name__ == "__main__":
    main()
